import streamlit as st
import cv2
import numpy as np
import os
import io
import sys
import time
import tempfile
import subprocess
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Detect if running on Streamlit Cloud (or any headless server) ─────────────
def is_cloud() -> bool:
    """True when running on Streamlit Cloud / any headless environment."""
    return (
        os.environ.get("STREAMLIT_SHARING_MODE") == "1"
        or os.environ.get("IS_STREAMLIT_CLOUD") == "1"
        or "/mount/src/" in os.getcwd()            # Streamlit Cloud path
        or not sys.stdin.isatty()                  # headless (no terminal attached)
    )

# ── Folder picker — local only ────────────────────────────────────────────────
def pick_folder() -> str:
    """Open OS native folder dialog. Returns '' if unavailable (headless/cloud)."""
    # Windows
    if sys.platform == "win32":
        try:
            ps = (
                "Add-Type -AssemblyName System.Windows.Forms;"
                "$d=New-Object System.Windows.Forms.FolderBrowserDialog;"
                "$d.Description='Select Video Folder';"
                "$d.RootFolder='MyComputer';"
                "if($d.ShowDialog() -eq 'OK'){Write-Output $d.SelectedPath}"
            )
            r = subprocess.run(
                ["powershell", "-NoProfile", "-NonInteractive", "-Command", ps],
                capture_output=True, text=True, timeout=60
            )
            return r.stdout.strip()
        except Exception:
            pass
    # macOS
    if sys.platform == "darwin":
        try:
            r = subprocess.run(
                ["osascript", "-e",
                 'POSIX path of (choose folder with prompt "Select Video Folder")'],
                capture_output=True, text=True, timeout=60
            )
            return r.stdout.strip().rstrip("/") if r.returncode == 0 else ""
        except Exception:
            pass
    # Linux / fallback tkinter
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk(); root.withdraw()
        root.wm_attributes("-topmost", True)
        sel = filedialog.askdirectory(title="Select Video Folder")
        root.destroy(); return sel or ""
    except Exception:
        pass
    # zenity (GNOME)
    try:
        r = subprocess.run(
            ["zenity", "--file-selection", "--directory",
             "--title=Select Video Folder"],
            capture_output=True, text=True, timeout=60
        )
        return r.stdout.strip() if r.returncode == 0 else ""
    except Exception:
        return ""

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Video QC Tool",
    page_icon="🎬",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ── CSS ────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Mono:wght@400;700&family=DM+Sans:wght@300;400;500;600;700&display=swap');
:root{--bg:#0b0b14;--card:#13131f;--border:#2a2a45;--accent:#7c6fff;
      --accent2:#ff6b9d;--green:#00e5a0;--red:#ff4d6a;--yellow:#ffd166;
      --text:#e8e8f0;--muted:#7a7a9a;}
html,body,[data-testid="stAppViewContainer"]{background-color:var(--bg)!important;
  font-family:'DM Sans',sans-serif;color:var(--text);}
[data-testid="stAppViewContainer"]>.main{background-color:var(--bg)!important;}
[data-testid="block-container"]{padding:2rem 3rem!important;max-width:1100px!important;margin:0 auto;}
#MainMenu,footer,header{visibility:hidden;}
[data-testid="stToolbar"]{display:none;}
[data-testid="stTextInput"] input,[data-testid="stNumberInput"] input{
  background:var(--card)!important;border:1px solid var(--border)!important;
  color:var(--text)!important;border-radius:8px!important;font-family:'DM Sans',sans-serif!important;}
[data-testid="stTextInput"] input:focus,[data-testid="stNumberInput"] input:focus{
  border-color:var(--accent)!important;box-shadow:0 0 0 2px rgba(124,111,255,0.2)!important;}
[data-testid="stSlider"]>div>div>div{background:var(--accent)!important;}
[data-testid="stButton"]>button{background:var(--accent)!important;color:white!important;
  border:none!important;border-radius:10px!important;font-family:'DM Sans',sans-serif!important;
  font-weight:600!important;font-size:15px!important;padding:0.5rem 1.5rem!important;transition:all .2s!important;}
[data-testid="stButton"]>button:hover{transform:translateY(-1px)!important;
  box-shadow:0 6px 24px rgba(124,111,255,0.4)!important;}
[data-testid="stButton"]>button:disabled{opacity:.4!important;transform:none!important;}
[data-testid="stProgress"]>div>div{
  background:linear-gradient(90deg,var(--accent),var(--accent2))!important;border-radius:4px!important;}
hr{border-color:var(--border)!important;}
label,[data-testid="stWidgetLabel"]{color:var(--muted)!important;font-size:13px!important;
  font-weight:500!important;letter-spacing:.4px!important;}
[data-testid="stDownloadButton"]>button{background:#1a3a2a!important;
  border:1px solid var(--green)!important;color:var(--green)!important;
  border-radius:10px!important;font-weight:600!important;}
[data-testid="stDownloadButton"]>button:hover{background:rgba(0,229,160,0.15)!important;
  transform:translateY(-1px)!important;}
[data-testid="stMetric"]{background:var(--card)!important;border:1px solid var(--border)!important;
  border-radius:12px!important;padding:1rem 1.2rem!important;}
[data-testid="stMetric"] label{font-size:12px!important;letter-spacing:1px!important;text-transform:uppercase!important;}
[data-testid="stMetricValue"]{font-family:'Space Mono',monospace!important;font-size:32px!important;font-weight:700!important;}
.stAlert{border-radius:10px!important;}
/* Mode toggle tabs */
div[data-testid="stTabs"] button[role="tab"]{
  font-family:'DM Sans',sans-serif!important;font-weight:600!important;font-size:13px!important;
  letter-spacing:.5px!important;}
/* File uploader */
[data-testid="stFileUploader"]{background:var(--card)!important;border:1px dashed var(--border)!important;
  border-radius:12px!important;padding:8px!important;}
[data-testid="stFileUploader"]:hover{border-color:var(--accent)!important;}
</style>
""", unsafe_allow_html=True)

# ── Constants ──────────────────────────────────────────────────────────────────
VIDEO_EXTS = (".mp4", ".mov", ".mkv", ".avi", ".wmv", ".flv", ".webm")
MAX_WORKERS = 4

# ── Core QC ───────────────────────────────────────────────────────────────────
def check_video(path: str, exp_w: int, exp_h: int, n_samples: int,
                border_px: int = 15, threshold: int = 10) -> dict:
    fname = os.path.basename(path)
    file_size_mb = round(os.path.getsize(path) / 1024 / 1024, 1)
    cap = cv2.VideoCapture(path)
    if not cap.isOpened():
        return _error_row(fname, path, file_size_mb, "Cannot open file")
    w   = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    h   = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    fps = cap.get(cv2.CAP_PROP_FPS)
    tot = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    dur = round(tot / fps, 1) if fps > 0 else "N/A"
    res_ok = (w == exp_w and h == exp_h)
    black = False
    sample_indices = np.linspace(0, max(tot - 1, 0), min(n_samples, max(tot, 1)), dtype=int)
    for idx in sample_indices:
        cap.set(cv2.CAP_PROP_POS_FRAMES, int(idx))
        ret, frame = cap.read()
        if not ret:
            continue
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        edges = [gray[:border_px,:], gray[-border_px:,:],
                 gray[:,:border_px], gray[:,-border_px:]]
        if any(np.mean(e) < threshold for e in edges):
            black = True; break
        del frame, gray
    cap.release()
    passed = res_ok and not black
    return {
        "File Name":         fname,
        "Full Path":         path,
        "File Size (MB)":    file_size_mb,
        "Status":            "PASS" if passed else "FAIL",
        "Resolution OK":     "Yes" if res_ok else "No",
        "Actual Resolution": f"{w}x{h}",
        "Expected":          f"{exp_w}x{exp_h}",
        "Duration (s)":      dur,
        "FPS":               round(fps, 2),
        "Black Border":      "Yes" if black else "No",
        "Checked At":        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

def check_video_from_bytes(fname: str, data: bytes, exp_w: int, exp_h: int,
                           n_samples: int) -> dict:
    """Write uploaded bytes to a temp file, run QC, then delete."""
    suffix = os.path.splitext(fname)[1] or ".mp4"
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp.write(data); tmp_path = tmp.name
    try:
        result = check_video(tmp_path, exp_w, exp_h, n_samples)
        result["File Name"] = fname          # restore original name
        result["Full Path"] = fname
        return result
    finally:
        try: os.unlink(tmp_path)
        except Exception: pass

def _error_row(fname, path, size_mb, reason):
    return {
        "File Name": fname, "Full Path": path, "File Size (MB)": size_mb,
        "Status": "FAIL", "Resolution OK": "ERROR",
        "Actual Resolution": reason, "Expected": "—",
        "Duration (s)": "ERROR", "FPS": "ERROR",
        "Black Border": "ERROR",
        "Checked At": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

def scan_folder(folder_path: str) -> list:
    try: entries = os.listdir(folder_path)
    except PermissionError: return []
    paths = []
    for f in sorted(entries):
        if f.lower().endswith(VIDEO_EXTS):
            full = os.path.join(folder_path, f)
            if os.path.isfile(full): paths.append(full)
    return paths

def fmt_size(mb):
    if isinstance(mb, str): return mb
    return f"{mb} MB" if mb < 1024 else f"{mb/1024:.1f} GB"

# ── Export ─────────────────────────────────────────────────────────────────────
EXPORT_COLS = [
    "File Name", "File Size (MB)", "Status", "Resolution OK",
    "Actual Resolution", "Expected", "Duration (s)", "FPS",
    "Black Border", "Checked At"
]

def to_excel(results: list) -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = "QC Report"
    th = Side(style="thin", color="2d2d4e")
    bd = Border(left=th, right=th, top=th, bottom=th)
    hf = PatternFill("solid", fgColor="1a237e")
    pf = PatternFill("solid", fgColor="1b5e20")
    ff = PatternFill("solid", fgColor="b71c1c")
    af = PatternFill("solid", fgColor="1a1a2e")
    for c, h in enumerate(EXPORT_COLS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hf
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = bd
    ws.row_dimensions[1].height = 28
    for ri, rec in enumerate(results, 2):
        ip = rec.get("Status") == "PASS"
        for ci, key in enumerate(EXPORT_COLS, 1):
            val = rec.get(key, "")
            cell = ws.cell(row=ri, column=ci, value=str(val))
            cell.border = bd
            cell.alignment = Alignment(
                horizontal="left" if ci == 1 else "center", vertical="center")
            if key == "Status":
                cell.fill = pf if ip else ff
                cell.font = Font(bold=True, color="FFFFFF", size=11)
            elif ri % 2 == 0:
                cell.fill = af
        ws.row_dimensions[ri].height = 22
    col_widths = [38, 14, 10, 14, 18, 14, 14, 8, 14, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

# ── Table HTML ─────────────────────────────────────────────────────────────────
def results_table_html(results: list) -> str:
    rows = ""
    for i, r in enumerate(results):
        sc = "#00e5a0" if r["Status"] == "PASS" else "#ff4d6a"
        si = "✅ PASS"  if r["Status"] == "PASS" else "❌ FAIL"
        rc = "#00e5a0" if r["Resolution OK"] == "Yes" else "#ff4d6a"
        bc = "#ff4d6a" if r["Black Border"]  == "Yes" else "#00e5a0"
        bg = "#13131f" if i % 2 == 0 else "#0f0f1a"
        size_str = fmt_size(r.get("File Size (MB)", "—"))
        rows += f"""
        <tr style="background:{bg}">
          <td style="padding:10px 14px;color:#e8e8f0;font-size:13px;max-width:280px;
              overflow:hidden;text-overflow:ellipsis;white-space:nowrap;
              font-family:'DM Sans',sans-serif" title="{r['File Name']}">{r['File Name']}</td>
          <td style="padding:10px 14px;text-align:center;color:#7a7a9a;
              font-size:12px;font-family:'Space Mono',monospace">{size_str}</td>
          <td style="padding:10px 14px;text-align:center">
            <span style="color:{sc};font-weight:700;font-size:13px;
                font-family:'Space Mono',monospace">{si}</span></td>
          <td style="padding:10px 14px;text-align:center;color:{rc};
              font-size:13px;font-family:'Space Mono',monospace">{r['Actual Resolution']}</td>
          <td style="padding:10px 14px;text-align:center;color:#7a7a9a;
              font-size:13px;font-family:'Space Mono',monospace">{r['Duration (s)']}s</td>
          <td style="padding:10px 14px;text-align:center;color:#7a7a9a;
              font-size:12px;font-family:'Space Mono',monospace">{r.get('FPS','—')}</td>
          <td style="padding:10px 14px;text-align:center;color:{bc};font-size:13px">
            {"⚠ Yes" if r["Black Border"] == "Yes" else "✓ No"}</td>
        </tr>"""
    return f"""
    <div style="border:1px solid #2a2a45;border-radius:14px;overflow:hidden;
                background:#13131f;margin-top:8px">
      <div style="padding:14px 20px;border-bottom:1px solid #2a2a45;
                  color:#7c6fff;font-weight:600;font-size:12px;
                  letter-spacing:1.5px;font-family:'DM Sans',sans-serif">
        📋 RESULTS
      </div>
      <div style="overflow-x:auto">
        <table style="width:100%;border-collapse:collapse">
          <thead><tr style="background:#0d0d1a">
            {''.join(f'<th style="padding:10px 14px;text-align:{"left" if i==0 else "center"};color:#7a7a9a;font-size:10px;letter-spacing:1.5px;font-family:DM Sans,sans-serif;font-weight:600">{h}</th>'
              for i,h in enumerate(["FILE NAME","SIZE","STATUS","RESOLUTION","DURATION","FPS","BLACK BORDER"]))}
          </tr></thead>
          <tbody>{rows}</tbody>
        </table>
      </div>
    </div>"""

# ── Session state ──────────────────────────────────────────────────────────────
for k, v in [("results", []), ("running", False), ("cancelled", False),
             ("folder_path", ""), ("input_mode", None)]:
    if k not in st.session_state:
        st.session_state[k] = v

# ── Header ─────────────────────────────────────────────────────────────────────
st.markdown("""
<div style="text-align:center;padding:36px 0 16px">
  <div style="font-size:48px;margin-bottom:8px">🎬</div>
  <h1 style="font-family:'Space Mono',monospace;color:#e8e8f0;
             font-size:28px;font-weight:700;letter-spacing:2px;margin:0">
    VIDEO QC TOOL
  </h1>
  <p style="color:#7a7a9a;font-size:13px;margin:8px 0 0;font-family:'DM Sans',sans-serif">
    Local folder · Upload files · 1 GB files supported · Parallel processing
  </p>
</div>
<hr style="border-color:#2a2a45;margin:0 0 20px">
""", unsafe_allow_html=True)

# ── Mode selector ──────────────────────────────────────────────────────────────
CLOUD = is_cloud()

# Show info banner on cloud
if CLOUD:
    st.markdown(
        "<div style='background:#0d1a2e;border:1px solid #1a3a6e;border-radius:10px;"
        "padding:10px 16px;margin-bottom:16px;font-size:13px;color:#7ab8ff'>"
        "☁️ <b>Running on Streamlit Cloud</b> — local folder browsing is unavailable. "
        "Use <b>Upload Files</b> mode, or run <code>streamlit run app.py</code> locally "
        "to use the Local Folder + Browse button.</div>",
        unsafe_allow_html=True
    )

tab_upload, tab_local = st.tabs(["☁️  Upload Files", "💻  Local Folder"])

video_files   = []          # list[str] — paths on disk, used by Run QC
uploaded_mode = False       # True when using temp files from uploader

# ── Tab 1: Upload ──────────────────────────────────────────────────────────────
with tab_upload:
    st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)
    st.markdown(
        "<div style='color:#7a7a9a;font-size:12px;margin-bottom:8px'>"
        "Select one or more video files from your computer. Works on Streamlit Cloud and locally.</div>",
        unsafe_allow_html=True
    )
    uploaded = st.file_uploader(
        "upload_videos",
        type=["mp4","mov","mkv","avi","wmv","flv","webm"],
        accept_multiple_files=True,
        label_visibility="collapsed",
        help="Select multiple files with Ctrl+Click / Cmd+Click",
        key="file_uploader"
    )
    if uploaded:
        total_upload_mb = sum(f.size for f in uploaded) / 1024 / 1024
        size_str = f"{total_upload_mb/1024:.1f} GB" if total_upload_mb >= 1024 else f"{total_upload_mb:.0f} MB"
        st.markdown(
            f"<div style='color:#00e5a0;font-size:13px;margin:6px 0'>"
            f"✅ <b>{len(uploaded)}</b> file(s) selected — total: <b>{size_str}</b></div>",
            unsafe_allow_html=True
        )
        with st.expander(f"📂 View file list ({len(uploaded)} files)", expanded=False):
            for f in uploaded:
                st.markdown(
                    f"<span style='font-family:Space Mono,monospace;font-size:12px;color:#7a7a9a'>"
                    f"{fmt_size(round(f.size/1024/1024,1))}  </span>"
                    f"<span style='font-size:13px;color:#e8e8f0'>{f.name}</span>",
                    unsafe_allow_html=True
                )
        # Store uploaded file objects in session for use by Run QC
        st.session_state["uploaded_files"] = uploaded
        st.session_state["input_mode"] = "upload"
    else:
        st.session_state["uploaded_files"] = []
        if st.session_state.get("input_mode") == "upload":
            st.session_state["input_mode"] = None

# ── Tab 2: Local Folder ────────────────────────────────────────────────────────
with tab_local:
    st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)
    if CLOUD:
        st.markdown(
            "<div style='color:#ffd166;font-size:13px;padding:10px 14px;"
            "background:#1f1a0a;border:1px solid #4a3800;border-radius:8px;margin-bottom:10px'>"
            "⚠️ Folder browsing requires running the app <b>locally</b>.<br>"
            "<code style='font-size:12px'>pip install -r requirements.txt<br>"
            "streamlit run app.py</code></div>",
            unsafe_allow_html=True
        )
    st.markdown(
        "<div style='color:#7a7a9a;font-size:13px;font-weight:500;"
        "letter-spacing:.4px;margin-bottom:6px'>📁  LOCAL FOLDER PATH</div>",
        unsafe_allow_html=True
    )
    col_input, col_browse = st.columns([5, 1])
    with col_input:
        typed = st.text_input(
            "folder_path_input",
            value=st.session_state.folder_path,
            placeholder="Windows: C:\\Users\\You\\Videos    |    Mac/Linux: /home/you/videos",
            label_visibility="collapsed",
            key="folder_text_input",
            disabled=CLOUD
        )
        if not CLOUD and typed != st.session_state.folder_path:
            st.session_state.folder_path = typed
            st.session_state["input_mode"] = "local"

    with col_browse:
        st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
        if st.button("🗂  Browse", use_container_width=True,
                     help="Open OS folder picker (local only — not available on Streamlit Cloud)",
                     disabled=CLOUD):
            picked = pick_folder()
            if picked:
                st.session_state.folder_path = picked
                st.session_state["input_mode"] = "local"
                st.session_state.pop("browse_error", None)
                st.rerun()
            else:
                st.session_state["browse_error"] = True

    if st.session_state.get("browse_error"):
        st.markdown(
            "<div style='color:#ffd166;font-size:12px;margin:4px 0 6px;padding:8px 12px;"
            "background:#1f1a0a;border:1px solid #4a3800;border-radius:8px'>"
            "⚠️ Folder dialog could not open — please paste the path directly above.</div>",
            unsafe_allow_html=True
        )

    folder_path = st.session_state.folder_path
    if folder_path.strip() and not CLOUD:
        fp = folder_path.strip()
        if not os.path.isdir(fp):
            st.markdown(
                "<div style='color:#ff4d6a;font-size:13px;margin:4px 0 8px'>"
                "⚠️ Folder not found — check the path.</div>", unsafe_allow_html=True)
        else:
            local_files = scan_folder(fp)
            total_size  = sum(os.path.getsize(p) for p in local_files) / 1024 / 1024
            size_str    = f"{total_size/1024:.1f} GB" if total_size >= 1024 else f"{total_size:.0f} MB"
            if local_files:
                st.markdown(
                    f"<div style='color:#00e5a0;font-size:13px;margin:4px 0 8px'>"
                    f"✅ Found <b>{len(local_files)}</b> video file(s) — total: <b>{size_str}</b></div>",
                    unsafe_allow_html=True)
                with st.expander(f"📂 View file list ({len(local_files)} files)", expanded=False):
                    for vp in local_files:
                        sz = os.path.getsize(vp) / 1024 / 1024
                        st.markdown(
                            f"<span style='font-family:Space Mono,monospace;font-size:12px;"
                            f"color:#7a7a9a'>{fmt_size(round(sz,1))}  </span>"
                            f"<span style='font-size:13px;color:#e8e8f0'>{os.path.basename(vp)}</span>",
                            unsafe_allow_html=True)
                st.session_state["local_files"] = local_files
                st.session_state["input_mode"] = "local"
            else:
                st.markdown(
                    "<div style='color:#ffd166;font-size:13px;margin:4px 0 8px'>"
                    "⚠️ No video files found (.mp4 .mov .mkv .avi .wmv .flv .webm)</div>",
                    unsafe_allow_html=True)
                st.session_state["local_files"] = []
    else:
        st.session_state.setdefault("local_files", [])


# ── Progress helper ────────────────────────────────────────────────────────────
def _update_progress(prog, status_ph, live_ph, results, done_count, total, t_start, row):
    elapsed  = time.time() - t_start
    avg_time = elapsed / done_count
    remaining = avg_time * (total - done_count)
    eta_str  = f"{int(remaining//60)}m {int(remaining%60)}s" if remaining > 0 else "—"
    pct = done_count / total
    prog.progress(pct, text=f"Processing {done_count}/{total} — ETA {eta_str}")
    passed_so_far = sum(1 for r in results if r["Status"] == "PASS")
    failed_so_far = done_count - passed_so_far
    status_ph.markdown(
        f"<div style='font-size:13px;color:#7a7a9a;font-family:DM Sans,sans-serif;"
        f"margin-bottom:6px'>Last: <b style='color:#e8e8f0'>{row['File Name']}</b> "
        f"— <span style='color:{'#00e5a0' if row['Status']=="PASS" else '#ff4d6a'}'>"
        f"{row['Status']}</span> &nbsp;|&nbsp; "
        f"✅ {passed_so_far} passed &nbsp; ❌ {failed_so_far} failed</div>",
        unsafe_allow_html=True
    )
    sorted_so_far = sorted(results, key=lambda r: r["Status"])
    live_ph.markdown(results_table_html(sorted_so_far), unsafe_allow_html=True)

# ── Resolve video_files for Run QC ────────────────────────────────────────────
input_mode = st.session_state.get("input_mode")
if input_mode == "upload":
    # Count uploaded objects as "files" — actual temp writes happen during QC
    video_files = st.session_state.get("uploaded_files", [])
    uploaded_mode = True
elif input_mode == "local":
    video_files = st.session_state.get("local_files", [])
    uploaded_mode = False
else:
    video_files = []

st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)

# ── Settings ───────────────────────────────────────────────────────────────────
c1, c2, c3, c4 = st.columns(4)
with c1:
    exp_w = st.number_input("Expected Width (px)", value=1080, min_value=1, step=1)
with c2:
    exp_h = st.number_input("Expected Height (px)", value=1920, min_value=1, step=1)
with c3:
    n_samples = st.slider("Frame Samples", 5, 30, 10,
                          help="Frames checked per video. More = slower but thorough.")
with c4:
    workers = st.slider("Parallel Workers", 1, 8, MAX_WORKERS,
                        help="Videos processed simultaneously. Lower if RAM is tight.")

st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

# ── Buttons ────────────────────────────────────────────────────────────────────
col_run, col_stop, col_clear = st.columns([3, 1, 1])
with col_run:
    run_disabled = st.session_state.running or len(video_files) == 0
    run = st.button("▶  Run QC", use_container_width=True, disabled=run_disabled)
with col_stop:
    stop = st.button("⏹  Stop", use_container_width=True,
                     disabled=not st.session_state.running)
with col_clear:
    if st.button("✕  Clear", use_container_width=True):
        st.session_state.results   = []
        st.session_state.running   = False
        st.session_state.cancelled = False
        st.rerun()

if stop:
    st.session_state.cancelled = True

# ── Run QC ─────────────────────────────────────────────────────────────────────
if run and video_files and not st.session_state.running:
    st.session_state.running   = True
    st.session_state.cancelled = False
    st.session_state.results   = []

    total      = len(video_files)
    results    = []
    done_count = 0

    prog      = st.progress(0.0, text="Starting…")
    status_ph = st.empty()
    live_ph   = st.empty()
    t_start   = time.time()

    if uploaded_mode:
        # ── Upload mode: write each file to temp, check, delete ──────────────
        # Read all bytes first (must be done in main thread)
        file_data = [(f.name, f.read()) for f in video_files]

        with ThreadPoolExecutor(max_workers=workers) as executor:
            futures = {
                executor.submit(check_video_from_bytes,
                                fname, data, exp_w, exp_h, n_samples): fname
                for fname, data in file_data
            }
            for future in as_completed(futures):
                if st.session_state.cancelled:
                    executor.shutdown(wait=False, cancel_futures=True)
                    break
                try:
                    row = future.result()
                except Exception as e:
                    fname = futures[future]
                    row   = _error_row(fname, fname, 0, str(e))
                results.append(row)
                done_count += 1
                _update_progress(prog, status_ph, live_ph, results,
                                 done_count, total, t_start, row)
    else:
        # ── Local mode: stream directly from disk ─────────────────────────────
        with ThreadPoolExecutor(max_workers=workers) as executor:
            futures = {
                executor.submit(check_video, vp, exp_w, exp_h, n_samples): vp
                for vp in video_files
            }
            for future in as_completed(futures):
                if st.session_state.cancelled:
                    executor.shutdown(wait=False, cancel_futures=True)
                    break
                try:
                    row = future.result()
                except Exception as e:
                    vp  = futures[future]
                    row = _error_row(os.path.basename(vp), vp,
                                     round(os.path.getsize(vp)/1024/1024, 1), str(e))
                results.append(row)
                done_count += 1
                _update_progress(prog, status_ph, live_ph, results,
                                 done_count, total, t_start, row)

    prog.empty(); status_ph.empty(); live_ph.empty()
    st.session_state.results = sorted(results, key=lambda r: r["Status"])
    st.session_state.running = False
    elapsed_total = time.time() - t_start
    if st.session_state.cancelled:
        st.warning(f"⏹ Stopped after {done_count}/{total} videos ({int(elapsed_total)}s elapsed).")
    else:
        st.success(f"✅ Done — {total} videos in {int(elapsed_total)}s "
                   f"({elapsed_total/total:.1f}s avg)")
    st.rerun()

# ── Results ────────────────────────────────────────────────────────────────────
if st.session_state.results:
    results = st.session_state.results
    total   = len(results)
    passed  = sum(1 for r in results if r["Status"] == "PASS")
    failed  = total - passed

    m1, m2, m3, m4 = st.columns(4)
    with m1: st.metric("✅ Passed", passed)
    with m2: st.metric("❌ Failed", failed)
    with m3: st.metric("🎬 Total",  total)
    with m4:
        total_size = sum(r.get("File Size (MB)", 0) for r in results
                         if isinstance(r.get("File Size (MB)"), (int, float)))
        st.metric("💾 Total Size", fmt_size(round(total_size, 1)))

    st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
    tab_all, tab_fail, tab_pass = st.tabs([
        f"All ({total})", f"❌ Failed ({failed})", f"✅ Passed ({passed})"])
    with tab_all:
        st.markdown(results_table_html(results), unsafe_allow_html=True)
    with tab_fail:
        fail_list = [r for r in results if r["Status"] == "FAIL"]
        if fail_list: st.markdown(results_table_html(fail_list), unsafe_allow_html=True)
        else: st.success("🎉 No failures!")
    with tab_pass:
        pass_list = [r for r in results if r["Status"] == "PASS"]
        if pass_list: st.markdown(results_table_html(pass_list), unsafe_allow_html=True)
        else: st.info("No videos passed.")

    st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
    dl1, dl2 = st.columns(2)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    export_data = [{k: r[k] for k in EXPORT_COLS if k in r} for r in results]
    with dl1:
        st.download_button("⬇️  Download CSV",
            data=pd.DataFrame(export_data).to_csv(index=False).encode(),
            file_name=f"video_qc_{ts}.csv", mime="text/csv",
            use_container_width=True)
    with dl2:
        st.download_button("⬇️  Download Excel",
            data=to_excel(export_data),
            file_name=f"video_qc_{ts}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)
